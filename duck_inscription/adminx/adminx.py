# coding=utf-8
from __future__ import unicode_literals
import datetime
from django.contrib.auth.admin import csrf_protect_m
from django.template.response import TemplateResponse
from django.views.decorators.cache import never_cache

from openpyxl.writer.excel import save_virtual_workbook
from duck_inscription.xadmin_plugins.topnav import IEDPlugin

from crispy_forms.bootstrap import TabHolder, Tab

from django.core.urlresolvers import reverse
from django.http import HttpResponseRedirect, HttpResponse
from mailrobot.models import Mail, MailBody, Address, Signature
from django.conf import settings
from xadmin.plugins.auth import UserAdmin

from xadmin.layout import Main, Fieldset, Container, Side, Row
from xadmin import views
import xadmin
from duck_inscription.models import Individu, SettingsEtape, WishWorkflow, SettingAnneeUni, WishParcourTransitionLog, \
    NoteMasterModel
from duck_inscription.models import Wish, SuiviDossierWorkflow, IndividuWorkflow, SettingsUser, CursusEtape
from xadmin.util import User
from xadmin.views import filter_hook, CommAdminView, BaseAdminView
from openpyxl import Workbook
import xlwt


class IncriptionDashBoard(views.website.IndexView):
    widgets = [[{"type": "qbutton", "title": "Gestion dossier",
                 "btns": [{'title': 'Reception', 'url': '/dossier_receptionner'},
                          {'title': 'Gestion Equivalence', 'url': '/dossier_equivalence'},
                          {'title': 'Gestion Dossier inscription', 'url': '/traitement_inscription'},
                          {'title': 'Remontee opi', 'model': Wish},
                          ]},
                {"type": "qbutton", "title": "Consultation des dossiers",
                "btns": [{'title': 'Consultation dossier inscription', 'model': Individu}, ]}
               ]]
    site_title = 'Backoffice'
    title = 'Accueil'
    widget_customiz = False


xadmin.site.register_view(r'^inscription/$', IncriptionDashBoard, 'inscription')


class StatistiqueDashBoard(views.website.IndexView):
    widgets = [[{"type": "qbutton", "title": "Inscription",
                 "btns": [{'title': 'Statistique Pal (équivalence, candidature)', 'url': '/stats_pal'},
                          {'title': 'Statistique Piel (préinscription)', 'url': '/stats_piel'}, ]}, ]]
    site_title = 'Backoffice'
    title = 'Accueil'
    widget_customiz = False


xadmin.site.register_view(r'^statistiques/$', StatistiqueDashBoard, 'statistiques')


class StatistiquePal(views.Dashboard):
    base_template = 'statistique/stats_pal.html'
    widget_customiz = False

    def get_context(self):
        context = super(StatistiquePal, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_pal'), 'title': 'Statistique PAL'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())


xadmin.site.register_view(r'^stats_pal/$', StatistiquePal, 'stats_pal')


class StatistiquePiel(views.Dashboard):
    base_template = 'statistique/stats_piel.html'
    widget_customiz = False

    @filter_hook
    def get_context(self):
        context = super(StatistiquePiel, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_piel'), 'title': 'Statistique PIEL'}]


xadmin.site.register_view(r'^stats_piel/$', StatistiquePiel, 'stats_piel')

class ExtractionPiel(views.Dashboard):
    base_template = 'extraction/extraction_pal.html'
    widget_customiz = False

    @filter_hook
    def get_context(self):
        context = super(ExtractionPiel, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_piel'), 'title': 'Statistique PIEL'}]


xadmin.site.register_view(r'^extraction/$', ExtractionPiel, 'extraction')


class ExtrationPalView(BaseAdminView):
    def get(self, request, *args, **kwargs):
        cod_etp = kwargs.get('step', '')
        step = SettingsEtape.objects.get(cod_etp=cod_etp)
        wb = xlwt.Workbook()
        ws = wb.add_sheet('etudiant')
        queryset = step.wish_set.filter(annee__cod_anu=2014, is_reins=False).order_by('individu__last_name')
        ws.write(1, 0, "code etudiant")
        ws.write(1, 1, "nom")
        ws.write(1, 2, "prenom")
        ws.write(1, 3, "code_dossier")
        ws.write(1, 4, "email")
        ws.write(1, 5, "moyenne generale")
        ws.write(1, 6, "note memoire")
        ws.write(1, 7, "note stage")
        for row, wish in enumerate(queryset):
            ws.write(row+2, 0, wish.individu.student_code)
            ws.write(row + 2, 1, wish.individu.last_name)
            ws.write(row + 2, 2, wish.individu.first_name1)
            ws.write(row + 2, 3, wish.code_dossier)
            ws.write(row + 2, 4, wish.individu.personal_email)
            try:
                ws.write(row + 2, 5, wish.notemastermodel.moyenne_general)
                ws.write(row + 2, 6, wish.notemastermodel.note_memoire)
                ws.write(row + 2, 7, wish.notemastermodel.note_stage)
            except NoteMasterModel.DoesNotExist:
                pass

        response = HttpResponse(mimetype='application/vnd.ms-excel')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xls' % ('extraction', date)
        wb.save(response)
        return response

xadmin.site.register_view(r'^extraction_note_view/(?P<step>\w+)/$', ExtrationPalView, 'extraction_note')

class MainDashboard(object):
    widgets = [[{"type": "qbutton", "title": "Scolarité", "btns": [

        {'title': "Pré-Inscription", 'url': 'inscription'}, {'title': "Statistique", 'url': 'statistiques'},
        {'title': 'Extraction', 'url': 'extraction'}]}, ]]
    site_title = 'Backoffice'
    title = 'Accueil'
    widget_customiz = False


xadmin.site.register(views.website.IndexView, MainDashboard)


class BaseSetting(object):
    use_bootswatch = True


xadmin.site.register(views.BaseAdminView, BaseSetting)


class GlobalSetting(object):
    menu_style = 'accordion'
    global_search_models = [Individu]
    global_add_models = []


xadmin.site.register(views.CommAdminView, GlobalSetting)


class WishInline(object):
    def email(self, obj):
        return obj.individu.personal_email

    def reins(self, obj):
        if obj.is_reins:
            return 'oui'
        else:
            return 'non'

    reins.short_description = 'Réinscription'

    model = Wish
    extra = 0
    style = 'table'
    fields = ['email', 'annee']
    readonly_fields = ['code_dossier', 'diplome_acces', 'centre_gestion', 'reins', 'date_validation', 'valide',
                       'get_transition_log', 'get_suivi_dossier', 'print_dossier_equi']
    exclude = ('annee', 'is_reins')
    can_delete = True
    hidden_menu = True

    @filter_hook
    def get_readonly_fields(self):
        if self.request.user.is_superuser:
            return self.readonly_fields
        else:
            return self.readonly_fields + ['state', 'suivi_dossier']

    def get_transition_log(self, obj):
        reponse = '<table>'
        for transition in obj.parcours_dossier.all():
            reponse += '<tr><td>{}</td><td>{}</td></tr>'.format(WishWorkflow.states[transition.to_state].title,
                                                                transition.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
        reponse += '</table>'
        return reponse

    get_transition_log.short_description = 'parcours'
    get_transition_log.allow_tags = True

    def get_suivi_dossier(self, obj):
        reponse = '<table>'
        print obj.etape_dossier.all()
        for transition in obj.etape_dossier.all():
            reponse += '<tr><td>{}</td><td>{}</td></tr>'.format(SuiviDossierWorkflow.states[transition.to_state].title,
                                                                transition.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
        reponse += '</table>'
        return reponse

    get_suivi_dossier.short_description = 'suivi'
    get_suivi_dossier.allow_tags = True

    def print_dossier_equi(self, obj):
        url = reverse('impression_equivalence', kwargs={'pk': obj.pk})
        url2 = reverse('impression_decision_equivalence', kwargs={'pk': obj.pk})
        reponse = '<a href="{}" class="btn btn-primary">Impression</a>'.format(url)
        reponse += '<a href="{}" class="btn btn-primary">ImpressionDecision</a>'.format(url2)
        return reponse

    print_dossier_equi.allow_tags = True
    print_dossier_equi.short_description = 'Impression dossier équivalence'


class IndividuXadmin(object):
    site_title = 'Consultation des dossiers étudiants'
    show_bookmarks = False
    fields = ('code_opi', 'last_name', 'first_name1', 'birthday', 'personal_email', 'state', 'user')
    readonly_fields = (
        'user', 'student_code', 'code_opi', 'last_name', 'first_name1', 'birthday', 'personal_email',
        'get_transition_log')
    list_display = ('__unicode__', 'last_name')
    list_per_page = 10
    search_fields = ('last_name', 'first_name1', 'common_name', 'student_code', 'code_opi', 'wishes__code_dossier')
    list_exclude = ('id', 'personal_email_save', 'opi_save', 'year')
    list_select_related = None
    inlines = [WishInline]
    hidden_menu = True

    def has_add_permission(self):
        return False

    def has_delete_permission(self, obj=None):
        return False

    @filter_hook
    def get_readonly_fields(self):
        if self.request.user.is_superuser:
            return self.readonly_fields
        else:
            return self.readonly_fields + ('state', )

    def get_transition_log(self, obj):
        reponse = '<table>'
        for transition in obj.etape_dossier.all():
            reponse += '<tr><td>{}</td><td>{}</td></tr>'.format(IndividuWorkflow.states[transition.to_state],
                                                                transition.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
        reponse += '</table>'
        return reponse

    get_transition_log.short_description = 'parcours'
    get_transition_log.allow_tags = True


class SettingsEtapeXadmin(object):
    exclude = ('lib_etp', 'cod_cyc', 'cod_cur', 'annee')
    list_display = ['__unicode__', 'date_ouverture_inscription', 'date_fermeture_inscription',
                    'date_fermeture_reinscription', 'droit', 'frais']
    list_filter = ['cursus']
    quickfilter = ['cursus']
    form_layout = (
        Main(Fieldset('Etape', 'cod_etp', 'diplome', 'cursus', 'label', 'label_formation'),
             TabHolder(Tab('Equivalence',
                                                                                                      Fieldset('',
                                                                                                   'date_ouverture_equivalence',
                                                                                                   'date_fermeture_equivalence',
                                                                                                   'document_equivalence',
                                                                                                   'path_template_equivalence',
                                                                                                   'grille_de_equivalence')),
                                                                                              Tab('Candidature',
                                                                                                  Fieldset('',
                                                                                                           'date_ouverture_candidature',
                                                                                                           'date_fermeture_candidature',
                                                                                                           'note_maste',
                                                                                                           'document_candidature', )), )),
        Side(Fieldset('Settings', 'required_equivalence', 'is_inscription_ouverte'))
    )


class UserSettingsInline(object):
    model = SettingsUser
    style_fields = {'etapes': 'm2m_transfer'}
    can_delete = False
    extra = 1
    max_num = 1


class CustomUserAdmin(UserAdmin):
    inlines = [UserSettingsInline]


class ExtrationStatistique(BaseAdminView):
    def get(self, request, *args, **kwargs):
        type_stat = kwargs.get('type_stat', 'stat_parcours_dossier')

        wb = Workbook()
        ws = wb.active

        if type_stat == 'stat_parcours_dossier':
            queryset = WishParcourTransitionLog.objects.filter(to_state=kwargs['etat'],
                                                               wish__etape__cod_etp=kwargs['step'])

            for row, x in enumerate(queryset):
                ws.cell(row + 1, 1).value = 'couou'

        else:
            queryset = Wish.objects.filter(state=kwargs['etat'], etape__cod_etp=kwargs['step'])
            for row, x in enumerate(queryset):
                ws.cell(row + 1, 1).value = 'couou'

        response = HttpResponse(save_virtual_workbook(wb), mimetype='application/vnd.ms-excel')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xls' % ('extraction', date)
        return response


class OpiView(object):
    model = Wish
    show_bookmarks = False
    list_export = []
    list_per_page = 10
    search_fields = ['code_dossier']
    # list_display_links_details = True
    # hidden_menu = True
    list_display = ('__str__', 'opi_url')

    def opi_url(self, obj):
        if obj.state.is_inscription:
            return '<a class="btn btn-primary" href="?opi={}">Remontée Opi</a>'.format(obj.code_dossier)
        else:
            return ''
    opi_url.short_description = 'remontee opi'
    opi_url.allow_tags = True
    opi_url.is_column = True

    def has_delete_permission(self, obj=None):
        return False

    def has_add_permission(self):
        return False

    @csrf_protect_m
    @filter_hook
    def get(self, request, *args, **kwargs):
        """
        The 'change list' admin view for this model.
        """
        response = self.get_result_list()
        if response:
            return response

        context = self.get_context()
        context.update(kwargs or {})

        response = self.get_response(context, *args, **kwargs)
        opi = self.request.GET.get('opi', None)
        if opi:
            wish = Wish.objects.get(code_dossier=opi)
            wish.save_opi()
            self.message_user('Etudiant {} remontee'.format(wish.individu.code_opi), 'success')

        return response or TemplateResponse(request, self.object_list_template or
                                            self.get_template_list('views/model_list.html'), context, current_app=self.admin_site.name)

xadmin.site.register_view(r'^extraction/(?P<type_stat>\w+)/(?P<etat>\w+)/(?P<step>\w+)/$', ExtrationStatistique,
                          'extraction_stat')
xadmin.site.unregister(User)
xadmin.site.register(User, CustomUserAdmin)
xadmin.site.register(Individu, IndividuXadmin)
xadmin.site.register(SettingsEtape, SettingsEtapeXadmin)
xadmin.site.register_plugin(IEDPlugin, CommAdminView)
xadmin.site.register(MailBody)
xadmin.site.register(Mail)
xadmin.site.register(Address)
xadmin.site.register(Signature)
xadmin.site.register(CursusEtape)
xadmin.site.register(SettingAnneeUni)
xadmin.site.register(Wish, OpiView)
# xadmin.site.register_view(r'^opi/$', OpiView, 'opi')
